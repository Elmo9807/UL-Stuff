import openpyxl
# Replace bracket input with desired filename
dataframe = openpyxl.load_workbook("C:\Python Test Files\AttendanceSheet.xlsx")
dataframe1 = dataframe.active
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        print(col[row].value)
